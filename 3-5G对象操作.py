from openpyxl import load_workbook
from openpyxl import Workbook
import re
import copy

"""
将5G表格转换成需要的格式，然后存储成ecxel表格
"""


class GnrcChannel(object):
    wb1 = load_workbook('5GNRChannel办公.xlsx')
    wb2 = Workbook()

    def __init__(self):
        # 获取表格中所有表名字，是一个列表
        self.gnrc_sheets = self.wb1.sheetnames

    # 读取原始表格中的数据，并返回字典，
    def get_info_gnrc(self, sheet_name):
        sheet = self.wb1[sheet_name]
        # 输出表的最大行和最大列
        print(sheet.max_row)
        print(sheet.max_column)
        # 读取表中的数据Bandwidth [MHz],Carrier centre,
        # 读出每一行所有值
        col_dict = {'1': 'A', '2': 'B', '3': 'C', '4': 'D', '5': 'E', '6': 'F', '7': 'G', '8': 'H', '9': 'I', '10': 'J',
                    '11': 'K', '12': 'L', '13': 'M', '14': 'N', '15': 'O', '16': 'P', '17': 'Q', '18': 'R'}
        word_value_list = []  # 第一行对应每一列数据的列和值，元祖（列号和值）
        for i in sheet['1']:
            # print(i.row, i.column, i.value)
            word_value_list.append((col_dict[str(i.column)], i.value))
        print(word_value_list)
        # 获取指定列中有效数字的行和值列表，元祖（行号，值）
        num_value_list = []
        print(22222222222222)
        for col, val in word_value_list:
            if val in ['Bandwidth [MHz]']:
                for x in sheet[col]:
                    # 判断获取到的有效值是否是int，如果是None，将其赋值和上层行一样的值
                    # print(x.row, x.value)
                    if isinstance(x.value, int):
                        num_value_list.append((str(x.row), x.value))
                    elif x.value is None and num_value_list and isinstance(num_value_list[-1:][0][1], int):
                        num_value_list.append((str(x.row), num_value_list[-1:][0][1]))
        print(num_value_list)
        print(33333333333333)
        # 获取up系列的值，先判断是否有符号&，符号实在第三列C中,先获取到C列值，判断&是否在其中,在其中的就从之前的Down中获取up值
        c_value_list = []
        for i in sheet['C']:
            c_value_list.append(i.value)
        # 找到需要的数据，以列的形式保存到字典中，键是列中有效数字
        sheet_value_dict = []
        # 遍历行,获取到Down系列的值
        for row, val in num_value_list:  # 里面每个元素表示数字一行系列，元祖（row行号，值）
            # 遍历列
            data_value_list = []  # 保存一系列数据中列的值
            for col, word in word_value_list:  # 每个元素表示一列，元祖（col字母列号和值）
                # 符号不在其中的话，说明有up列的数据，所以需要再获取up的数据，
                if sheet[col + str(int(row))].value is None:
                    sheet[col + str(int(row))].value = sheet[col + str(int(row) - 1)].value
                    # 将获取到的行值保存到一个列表中，表示一行数据
                data_value_list.append(sheet[col + str(int(row))].value)
                # print(sheet[col + str(int(row))].value, end=",")  # 这里只表示一个单元格的值，遍历完才表示一行的数据
            if "&" in c_value_list:
                # 否则的话说明&在，需要将&替换成Downlink，并且遇到Uplink存在的行的时候，将Uplink替换成Downlink，然后后三项拼接到结尾，并将Downlink替换成Uplink
                if "&" in data_value_list:
                    data_value_list[2] = "Downlink"
                    sheet_value_dict.append(data_value_list)
                elif "Uplink" in data_value_list:
                    data_value_list[2] = "Downlink"
                    sheet_value_dict.append(data_value_list)
                    add_Uplink_list = copy.deepcopy(sheet_value_dict[-3:])
                    for x in add_Uplink_list:
                        x[2] = "Uplink"
                    sheet_value_dict.extend(add_Uplink_list)
                else:
                    sheet_value_dict.append(data_value_list)
            else:
                sheet_value_dict.append(data_value_list)
        # print(sheet_value_dict)
        # for value in sheet_value_dict:
        #     print(value)
        return sheet_value_dict

    # 清洗原始表格中的数据，返回需要的样式数据
    def wash_info_gnrc(self, sheet_name, sheet_value_list):
        # 获取band和scs和ul，dl的值，插入到每一行数据的开头
        re_com = re.compile(r'(\w\d*)_(\d*)kHz').match(sheet_name)
        band = re_com.group(1)
        scs = re_com.group(2)
        for row in sheet_value_list:
            ul_bandwidth = row[0]
            dl_bandwidth = row[0]
            row.insert(0, dl_bandwidth)
            # row.insert(0, ul_bandwidth)
            row.insert(0, scs)
            row.insert(0, band)
            print(row)
        return sheet_value_list

    # 保存成新表格样式表
    def save_new_sheet(self, sheet_name, clean_sheet_value):
        sheet = self.wb2.create_sheet(sheet_name)
        tit_list = ['band ', 'scs', 'ul-bandwidth [MHz]', 'dl-bandwidth [MHz]', 'carrierBandwidth[PRBs]', 'Duplex',
                    'FreqPos', 'Carrier centre [MHz]', 'Carrier centre [ARFCN]', 'point A [MHz]',
                    'absoluteFrequencyPointA [ARFCN]', 'offsetToCarrier [Carrier PRBs]', 'SS block SCS [kHz]', 'GSCN',
                    'absoluteFrequencySSB [ARFCN]', 'KSSB', 'CORESET#0 Offset [RBs]Note 1', 'CORESET#0 Index Note 1 ',
                    'offsetToPointA (SIB1) [PRBs] Note 1']
        sheet.append(tit_list)
        for row in clean_sheet_value:
            sheet.append(row)
        return sheet

    # 保存excel表格
    def save_excel(self, wb):
        wb.save("5G目标文件.xlsx")

    # 主程序
    def run(self):
        for sheet_name in self.gnrc_sheets:
            # 通过表名字获取指定表
            # sheet_name = "N38_15kHz"
            print(f'开始读取表{sheet_name}--------------------')
            # 获取原始表格信息
            sheet_value_list = self.get_info_gnrc(sheet_name)
            # 清洗数据
            clean_sheet_value = self.wash_info_gnrc(sheet_name, sheet_value_list)
            # 保存表信息到表中
            self.save_new_sheet(sheet_name, clean_sheet_value)
            # 保存表格到指定文件
            self.save_excel(self.wb2)
            print(f'读取结束了{sheet_name}--------------------')


if __name__ == '__main__':
    gnrc = GnrcChannel()
    print("----------程序开始执行------------")
    gnrc.run()
    print("----------程序执行结束------------")
